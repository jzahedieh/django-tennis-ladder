import os
from collections import defaultdict

from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook
from ladder.models import Season, Player, Ladder, League, LadderSubscription
import datetime


class Command(BaseCommand):
    help = 'Exports XLS, args: --season'

    def add_arguments(self, parser):
        parser.add_argument('--season',
                            action='store',
                            dest='season',
                            default=False,
                            help='ID of season')

        parser.add_argument('--file',
                            action='store',
                            dest='file',
                            default=False,
                            help='File location of import speadsheet')

    def handle(self, *args, **options):

        # make arg checks
        if options['file'] is False:
            raise CommandError('--file location is not set')

        file_location = options['file']

        if os.access(file_location, os.R_OK) is False:
            raise CommandError('Directory (%s) is not writeable, change in code' % file_location)

        # make arg checks
        if options['season'] is False:
            raise CommandError('--season id option not set')

        # book = open_workbook(file_location)
        book = load_workbook(file_location)
        season = Season.objects.get(pk=options['season'])
        sh1 = book['Sheet1']  # always first sheet
        player_list = defaultdict(dict)  # initialize defaultdict for our player list.
        current_div = {}  # set the division counter to 0
        subscriptions_created = 0

        # save all players then set up ladder
        for rows in sh1.iter_rows(values_only=True):

            # save division numbers and load ladder
            if rows[5] == 'DIVISION':
                current_div = rows[8]
                try:
                    ladder = Ladder.objects.get(season=season, division=current_div)
                except:
                    ladder = Ladder(season=season, division=current_div, ladder_type="First to 9")
                    ladder.save()

            # save players
            if rows[0] and rows[1] != 'NAME':
                position = rows[0]
                first_name = rows[1]
                last_name = rows[2]

                # populate our array for later use
                player_list[current_div][position] = {'first_name': first_name, 'last_name': last_name}

                try:
                    player_object = Player.objects.get(first_name=first_name, last_name=last_name)
                except:
                    print('No match for player: ' + first_name + ' ' + last_name)
                    try:
                        player_object = Player(first_name=first_name, last_name=last_name)
                        player_object.save()
                    except:
                        print('Failed to create player: ' + first_name + ' ' + last_name)

                try:
                    League.objects.get(ladder=ladder, player=player_object)
                except League.MultipleObjectsReturned:
                    pass
                except:
                    League.objects.create(ladder=ladder, player=player_object, sort_order=position * 10)

                # Auto-subscribe players who have linked user accounts
                subscription, created = LadderSubscription.objects.get_or_create(
                    ladder=ladder,
                    user=player_object.user,
                    defaults={'subscribed_at': datetime.date.today()}
                )
                if created:
                    subscriptions_created += 1

        print('built good')
        print(f'Created {subscriptions_created} new email subscriptions')
